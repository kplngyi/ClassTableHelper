import openpyxl
import re
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
# 课表地址
filename = '/Users/dy/Desktop/选课表带时间.xlsx'
# 打开Excel文件
workbook = openpyxl.load_workbook(filename)
sheet0 = workbook['sheet0']  # 或者选择您要读取的工作表
# 目标颜色的RGB值，这里是F4B084的RGB值
# target_color = "F4B08400"
# 测试是否读入文件
# print(sheet['D1'].value)
weekdays = [
    "周一",
    "周二",
    "周三",
    "周四",
    "周五",
    "周六",
    "周日"
]
sessions = [
    "1",
    "2",
    "3",
    "4",
    "5",
    "6",
    "7",
    "8",
    "9",
    "10",
    "11",
    "12"
]
time_info = {
    1: "8:30-9:20",
    2: "9:20-10:10",
    3: "10:30-11:20",
    4: "11:20-12:10",
    5: "13:30-14:20",
    6: "14:20-15:10",
    7: "15:30-16:20",
    8: "16:20-17:10",
    9: "18:10-19:00",
    10: "19:00-19:50",
    11: "20:10-21:00",
    12: "21:00-21:50"
}
# 创建一个新的工作簿和工作表--课表
classTable_workbook = openpyxl.Workbook()
classTable_sheet = classTable_workbook.active

# 在新工作表中写入列、开始行和结束行的值
# 初始化课表
classTable_sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)
classTable_sheet.cell(row=1, column=1, value="课表")
# 小节次序
for classTable_sheet_row in range(3, 15):
    classTable_sheet.cell(row=classTable_sheet_row, column=1, value=sessions[classTable_sheet_row - 3])
# 小节时间
for classTable_sheet_row in range(3, 15):
    classTable_sheet.cell(row=classTable_sheet_row, column=2, value=time_info[classTable_sheet_row - 2])
# 周
for classTable_sheet_col in range(3, 10):
    classTable_sheet.cell(row=2, column=classTable_sheet_col, value=weekdays[classTable_sheet_col - 3])


# 获取颜色
def get_color(sheet):
    target_color = ""
    for row in sheet.iter_rows():
        for cell in row:
            # 检查单元格的填充颜色是否与目标颜色匹配
            if cell.value == "投资学":
                target_color = cell.fill.start_color.rgb
    return target_color


# 将原始表坐标转化为新表坐标
def change_data(date_info, data_row, data_col):
    """
    根据日期信息、数据行和数据列计算新表格坐标

    :param date_info: xlsx文件第12列信息，例如："周一(3-4)"
    :param data_row: 数据行号
    :param data_col: 数据列号
    :return: 返回新表格的开始列、开始行和结束行
    """
    weekdays_dict = {
        "周一": 1,
        "周二": 2,
        "周三": 3,
        "周四": 4,
        "周五": 5,
        "周六": 6,
        "周日": 7
    }
    # 直接使用字符串在判断过程中出现问题，第6位有可能是'-'
    # 具体原因不清楚使用正则表达式找到节次信息
    pattern = r'\d+'
    sessions = re.findall(pattern, date_info)
    # 解析日期信息，提取开始列和结束行的数字
    start_col = data_col + weekdays_dict[date_info[:2]]
    start_row = data_row + int(sessions[0])
    end_row = data_row + int(sessions[1])

    return start_col, start_row, end_row


# 获取感兴趣课程的颜色
interest_color = get_color(sheet0)

# 打算获取 课程名称、课程属性、课时/学时、开课周、星期节次、教室、考核方式 7种信息
# columns_to_read = [4, 6, 8, 11, 12, 13, 15]
columns_to_read = [2, 3, 4, 8, 11, 12]
course_info = {
    "序号": 1,
    "开课院系": 2,
    "课程编码": 3,
    "课程名称": 4,
    "英文名称": 5,
    "课程属性": 6,
    "所属学科/专业": 7,
    "课时/学分": 8,
    "限选人数": 9,
    "已选人数": 10,
    "开课周": 11,
    "星期节次": 12,
    "教室": 13,
    "授课方式": 14,
    "考试方式": 15,
    "首席教授": 16,
    "首席教授所属单位": 17,
    "主讲教师": 18,
    "主讲教师所属单位": 19,
    "助教": 20,
    "助教所属单位": 21,
    "召集人": 22,
    "召集人所属单位": 23
}
time_info = {
    1: "8:30-9:20",
    2: "9:20-10:10",
    3: "10:30-11:20",
    4: "11:20-12:10",
    5: "13:30-14:20",
    6: "14:20-15:10",
    7: "15:30-16:20",
    8: "16:20-17:10",
    9: "18:10-19:00",
    10: "19:00-19:50",
    11: "20:10-21:00",
    12: "21:00-21:50"
}
# 互换后的字典
flipped_course_info = {value: key for key, value in course_info.items()}
# # 打印互换后的字典
# for key, value in flipped_course_info.items():
#     print(f"{key}: {value}")

# 冲突课程颜色提示
red_fill = PatternFill(start_color="5867B5", end_color="5867B5", fill_type="solid")
# 课程信息
classInfo = ""
# 读取全校大课表
# 遍历每一行并只处理指定列的数据
for row in sheet0.iter_rows(min_row=2):  # 第一行 是标题，从第二行开始遍历

    cell = row[0]  # 看颜色
    # 检查单元格颜色 ，如果是兴趣课，写入课表
    if cell.fill.start_color.rgb == interest_color:
        # 获得时间节次
        cell = row[11]
        time_data = cell.value
        # 获取需要写入数据的地址
        write_col, write_start_row, write_end_row = change_data(time_data, 2, 2)
        # 包含课程数据信息的数据列
        course_info = ""
        # 找到需要写入的数据列,记录课程信息
        for col in columns_to_read:
            cell = row[col - 1]
            cell_value = cell.value
            course_info = course_info + str(flipped_course_info[col]) + ":" + cell_value + '\n'
            if col == 3:
                classInfo = classInfo + str(flipped_course_info[col]) + ":" + cell_value + '\n'
            elif col == 4:
                classInfo = classInfo + str(flipped_course_info[col]) + ":" + cell_value + '\n'
            # course_info = course_info + cell_value + "\n"
        print(course_info)
        for row_index in range(write_start_row, write_end_row + 1):
            # 使用sheet.cell()方法获取单元格的值
            cell_value = classTable_sheet.cell(row=row_index, column=write_col).value
            # 检查单元格是否为空
            if cell_value is None or cell_value == "" or cell_value == course_info:
                classTable_sheet.cell(row=row_index, column=write_col, value=course_info)
            else:
                # 如果有数据，将单元格着色为红色
                classTable_sheet.cell(row=row_index, column=write_col).fill = red_fill
                classTable_sheet.cell(row=row_index, column=write_col, value=cell_value + course_info)

# 自动换行
# 创建 Alignment 对象并设置 wrap_text 为 True，以启用文本自动换行
alignment = Alignment(wrap_text=True)
# 遍历工作表中的所有单元格，并应用文本自动换行
for cell_row in classTable_sheet.iter_rows():
    for cell in cell_row:
        cell.alignment = alignment

# 设置所有行和全部列的长和宽
width = 120
height = 360
for i in range(1, classTable_sheet.max_row + 1):
    classTable_sheet.row_dimensions[i].height = height
for i in range(1, classTable_sheet.max_column + 1):
    classTable_sheet.column_dimensions[get_column_letter(i)].width = width

# 保存新工作簿
classTable_workbook.save('myClassTable.xlsx')

# 关闭excel文件
workbook.close()
print(classInfo)
